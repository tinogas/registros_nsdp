"""
Importa los archivos Excel de sacramentos a SQLite.

Maneja las diferencias entre ambos archivos:
- Filas de encabezados distintas por hoja
- Orden PAPA/MAMA invertido en primera_comunion del archivo 2023
- Nombre de hoja BAUTIZMO vs CONSTANCIA.BAUTISMO
- Columnas extra Dias/Mes en matrimonios 2023
"""
from pathlib import Path
from typing import Optional
import openpyxl
from app.core.database import db, init_db
from app.utils.config import EXCEL_FILES


def _cell(row, idx: int) -> Optional[str]:
    """Devuelve el valor de la celda como texto limpio o None."""
    if idx >= len(row):
        return None
    val = row[idx]
    if val is None:
        return None
    return str(val).strip() or None


def _detect_fuente(path: Path) -> str:
    name = path.name.upper()
    if "2023" in name:
        return "2023"
    return "2010-2022"


def _find_header_row(sheet, keywords: list[str]) -> int:
    """Busca la fila que contiene todos los keywords dados (case-insensitive)."""
    for i, row in enumerate(sheet.iter_rows(max_row=15, values_only=True)):
        row_text = " ".join(str(c).upper() for c in row if c)
        if all(k.upper() in row_text for k in keywords):
            return i  # índice 0-based
    return 0


# ---------------------------------------------------------------------------
# Importadores por tipo de sacramento
# ---------------------------------------------------------------------------

def _import_matrimonios(sheet, fuente: str, conn) -> int:
    is_2023 = fuente == "2023"
    header_row = _find_header_row(sheet, ["PAREJA", "PRESBITERO"])
    rows = list(sheet.iter_rows(min_row=header_row + 2, values_only=True))

    batch = []
    for row in rows:
        if not any(row):
            continue
        # Columnas: A=0 NUMERO, B=1 PAREJA, C=2 DIA, D=3 MES, E=4 AÑO,
        #           F=5 PRESBITERO, G=6 TESTIGO1, H=7 T2, I=8 T3, J=9 T4,
        #           K=10 PÁRROCO, L=11 LIBRO, M=12 PAGINA, N=13 PARTIDA
        #           (2023 agrega V=21 DIAS, W=22 MES)
        batch.append((
            _cell(row, 0),   # numero
            _cell(row, 1),   # pareja
            _cell(row, 2),   # dia
            _cell(row, 3),   # mes
            _cell(row, 4),   # anio
            _cell(row, 5),   # presbitero
            _cell(row, 6),   # testigo1
            _cell(row, 7),   # testigo2
            _cell(row, 8),   # testigo3
            _cell(row, 9),   # testigo4
            _cell(row, 10),  # parroco
            _cell(row, 11),  # libro
            _cell(row, 12),  # pagina
            _cell(row, 13),  # partida
            _cell(row, 21) if is_2023 else None,  # dias_extra
            _cell(row, 22) if is_2023 else None,  # mes_extra
            fuente,
        ))

    conn.executemany(
        """INSERT INTO matrimonios
           (numero,pareja,dia,mes,anio,presbitero,testigo1,testigo2,testigo3,testigo4,
            parroco,libro,pagina,partida,dias_extra,mes_extra,fuente_archivo)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        batch,
    )
    return len(batch)


def _import_primera_comunion(sheet, fuente: str, conn) -> int:
    is_2023 = fuente == "2023"
    header_row = _find_header_row(sheet, ["NIÑO", "PÁRROCO"])
    if header_row == 0:
        header_row = _find_header_row(sheet, ["NINO", "PARROCO"])
    rows = list(sheet.iter_rows(min_row=header_row + 2, values_only=True))

    batch = []
    for row in rows:
        if not any(row):
            continue
        # 2010-2022: C=2 NOMBRE, D=3 DIA, E=4 MES, F=5 AÑO, G=6 MAMA, H=7 PAPA, I=8 PADRINOS, J=9 PÁRROCO
        # 2023:      C=2 NOMBRE, D=3 DIA, E=4 MES, F=5 AÑO, G=6 PAPA, H=7 MAMA, I=8 PADRINOS, J=9 PÁRROCO
        if is_2023:
            mama_val = _cell(row, 7)
            papa_val = _cell(row, 6)
        else:
            mama_val = _cell(row, 6)
            papa_val = _cell(row, 7)

        batch.append((
            _cell(row, 2),   # nombre
            _cell(row, 3),   # dia
            _cell(row, 4),   # mes
            _cell(row, 5),   # anio
            mama_val,
            papa_val,
            _cell(row, 8),   # padrinos
            _cell(row, 9),   # parroco
            fuente,
        ))

    conn.executemany(
        """INSERT INTO primera_comunion
           (nombre,dia,mes,anio,mama,papa,padrinos,parroco,fuente_archivo)
           VALUES (?,?,?,?,?,?,?,?,?)""",
        batch,
    )
    return len(batch)


def _import_confirmacion(sheet, fuente: str, conn) -> int:
    header_row = _find_header_row(sheet, ["NIÑA", "ARZOBISPO"])
    if header_row == 0:
        header_row = _find_header_row(sheet, ["NINA", "ARZOBISPO"])
    rows = list(sheet.iter_rows(min_row=header_row + 2, values_only=True))

    batch = []
    for row in rows:
        if not any(row):
            continue
        # A=0 NUMERO, B=1 NOMBRE, C=2 DIA, D=3 MES, E=4 AÑO,
        # F=5 PAPA, G=6 MAMA, H=7 PADRINOS, I=8 ARZOBISPO, J=9 PARROCO,
        # K=10 LIBRO, L=11 PAGINA, M=12 PARTIDA
        batch.append((
            _cell(row, 0),
            _cell(row, 1),
            _cell(row, 2),
            _cell(row, 3),
            _cell(row, 4),
            _cell(row, 5),
            _cell(row, 6),
            _cell(row, 7),
            _cell(row, 8),
            _cell(row, 9),
            _cell(row, 10),
            _cell(row, 11),
            _cell(row, 12),
            fuente,
        ))

    conn.executemany(
        """INSERT INTO confirmacion
           (numero,nombre,dia,mes,anio,papa,mama,padrinos,arzobispo,parroco,libro,pagina,partida,fuente_archivo)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        batch,
    )
    return len(batch)


def _import_bautismos(sheet, fuente: str, conn) -> int:
    header_row = _find_header_row(sheet, ["NIÑA", "MINISTRO"])
    if header_row == 0:
        header_row = _find_header_row(sheet, ["NINA", "MINISTRO"])
    if header_row == 0:
        header_row = _find_header_row(sheet, ["BAUTISMO", "PADRINO"])
    rows = list(sheet.iter_rows(min_row=header_row + 2, values_only=True))

    batch = []
    for row in rows:
        if not any(row):
            continue
        # B=1 NOMBRE
        # D=3 DIA_NAC, E=4 MES_NAC, F=5 AÑO_NAC
        # G=6 LUGAR_NAC
        # H=7 PAPA, I=8 MAMA
        # J=9 DIA_BAU, K=10 MES_BAU, L=11 AÑO_BAU
        # M=12 MINISTRO
        # N=13 PADRINO, O=14 MADRINA
        # P=15 PÁRROCO
        # Q=16 REGISTRO_NO
        # U=20 LIBRO, V=21 PAGINA, W=22 ACTA
        batch.append((
            _cell(row, 1),
            _cell(row, 3),
            _cell(row, 4),
            _cell(row, 5),
            _cell(row, 6),
            _cell(row, 7),
            _cell(row, 8),
            _cell(row, 9),
            _cell(row, 10),
            _cell(row, 11),
            _cell(row, 12),
            _cell(row, 13),
            _cell(row, 14),
            _cell(row, 15),
            _cell(row, 16),
            _cell(row, 20),
            _cell(row, 21),
            _cell(row, 22),
            fuente,
        ))

    conn.executemany(
        """INSERT INTO bautismos
           (nombre,dia_nacimiento,mes_nacimiento,anio_nacimiento,lugar_nacimiento,
            papa,mama,dia_bautismo,mes_bautismo,anio_bautismo,ministro,padrino,madrina,
            parroco,registro_no,libro,pagina,acta,fuente_archivo)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        batch,
    )
    return len(batch)


def _import_catecumenos(sheet, fuente: str, conn) -> int:
    header_row = _find_header_row(sheet, ["NOMBRE", "PADRINOS"])
    rows = list(sheet.iter_rows(min_row=header_row + 2, values_only=True))

    batch = []
    for row in rows:
        if not any(row):
            continue
        # A=0 NOMBRE, B=1 DIA, C=2 MES, D=3 AÑO, E=4 PADRE, F=5 MADRE, G=6 PADRINOS
        batch.append((
            _cell(row, 0),
            _cell(row, 1),
            _cell(row, 2),
            _cell(row, 3),
            _cell(row, 4),
            _cell(row, 5),
            _cell(row, 6),
            fuente,
        ))

    conn.executemany(
        """INSERT INTO catecumenos (nombre,dia,mes,anio,padre,madre,padrinos,fuente_archivo)
           VALUES (?,?,?,?,?,?,?,?)""",
        batch,
    )
    return len(batch)


# Mapa: nombre de hoja normalizado → función importadora
_SHEET_MAP = {
    "MATRIMONIO":           _import_matrimonios,
    "1RA.COMUNION":         _import_primera_comunion,
    "1ERA.COMUNION":        _import_primera_comunion,
    "PRIMERA COMUNION":     _import_primera_comunion,
    "CONFIRMACION":         _import_confirmacion,
    "CONFIRMACIÓN":         _import_confirmacion,
    "CONSTANCIA.BAUTISMO":  _import_bautismos,
    "BAUTIZMO":             _import_bautismos,
    "BAUTISMO":             _import_bautismos,
    "CATECUMENOS":          _import_catecumenos,
    "CATECÚMENOS":          _import_catecumenos,
}


def import_excel(path: Path, progress_cb=None) -> dict:
    """
    Importa un archivo Excel a SQLite.
    progress_cb(mensaje: str) se llama opcionalmente con actualizaciones.
    Devuelve dict con conteos por tabla.
    """
    fuente = _detect_fuente(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    totals = {}

    with db() as conn:
        for sheet_name in wb.sheetnames:
            normalized = sheet_name.upper().replace("  ", " ").strip()
            fn = _SHEET_MAP.get(normalized)
            if fn is None:
                continue  # hoja de impresión u otra — saltar

            sheet = wb[sheet_name]
            if progress_cb:
                progress_cb(f"Importando: {sheet_name}...")
            try:
                count = fn(sheet, fuente, conn)
                totals[sheet_name] = count
                if progress_cb:
                    progress_cb(f"  -> {count} registros")
            except Exception as e:
                if progress_cb:
                    progress_cb(f"  ! Error en {sheet_name}: {e}")

    wb.close()
    return totals


def run_full_import(progress_cb=None) -> dict:
    """Inicializa la BD e importa todos los archivos Excel configurados."""
    from app.core.database import recalculate_all_folios
    init_db()
    all_totals = {}
    for path in EXCEL_FILES:
        if not path.exists():
            if progress_cb:
                progress_cb(f"Archivo no encontrado: {path.name}")
            continue
        if progress_cb:
            progress_cb(f"\n=== {path.name} ===")
        totals = import_excel(path, progress_cb)
        all_totals[path.name] = totals

    if progress_cb:
        progress_cb("\nCalculando folios por año...")
    recalculate_all_folios()
    if progress_cb:
        progress_cb("Folios asignados.")
    return all_totals


if __name__ == "__main__":
    result = run_full_import(print)
    print("\n--- Totales finales ---")
    total_global = 0
    for archivo, tablas in result.items():
        print(f"\n{archivo}:")
        for tabla, n in tablas.items():
            print(f"  {tabla}: {n}")
            total_global += n
    print(f"\nTotal importado: {total_global} registros")
