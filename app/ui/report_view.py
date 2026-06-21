"""
Módulo de reportes: filtros por tipo, año y mes → tabla de resultados
→ exportar a PDF o Excel.
"""
import datetime
import threading
import tempfile
from pathlib import Path
import tkinter as tk
from tkinter import ttk
import tkinter.font as tkfont
import customtkinter as ctk
from app.core.database import db
from app.ui.search_view import _style_treeview

MESES = [
    "Todos", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
    "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE",
]

TABLAS = {
    "Matrimonios":    "matrimonios",
    "1a Comunión":    "primera_comunion",
    "Confirmación":   "confirmacion",
    "Bautismos":      "bautismos",
    "Catecúmenos":    "catecumenos",
}

REPORT_COLS = {
    "matrimonios":     ["pareja",  "dia", "mes", "anio", "presbitero", "parroco"],
    "primera_comunion":["nombre",  "dia", "mes", "anio", "mama",       "papa",    "parroco"],
    "confirmacion":    ["nombre",  "dia", "mes", "anio", "arzobispo",  "parroco"],
    "bautismos":       ["nombre",  "dia_bautismo", "mes_bautismo", "anio_bautismo", "padrino", "madrina", "parroco"],
    "catecumenos":     ["nombre",  "dia", "mes", "anio", "padre", "madre"],
}

REPORT_HEADERS = {
    "pareja":        "Pareja",
    "nombre":        "Nombre",
    "dia":           "Día",
    "mes":           "Mes",
    "anio":          "Año",
    "presbitero":    "Presbítero",
    "parroco":       "Párroco",
    "mama":          "Mamá",
    "papa":          "Papá",
    "arzobispo":     "Arzobispo",
    "dia_bautismo":  "Día",
    "mes_bautismo":  "Mes",
    "anio_bautismo": "Año",
    "padrino":       "Padrino",
    "madrina":       "Madrina",
    "padre":         "Padre",
    "madre":         "Madre",
}

YEAR_COL = {
    "matrimonios":     "anio",
    "primera_comunion":"anio",
    "confirmacion":    "anio",
    "bautismos":       "anio_bautismo",
    "catecumenos":     "anio",
}

MES_COL = {
    "matrimonios":     "mes",
    "primera_comunion":"mes",
    "confirmacion":    "mes",
    "bautismos":       "mes_bautismo",
    "catecumenos":     "mes",
}

# Columnas numéricas / cortas (ancla centrada en treeview y mínimo estrecho en PDF)
_CENTER_COLS = {"dia", "dia_bautismo", "anio", "anio_bautismo"}


class ReportView(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Reportes")
        self.geometry("960x640")
        self.grab_set()
        self._results: list[dict] = []
        self._table = "bautismos"
        self._odd_row, self._even_row = _style_treeview()
        self._build()

    # ------------------------------------------------------------------
    def _build(self):
        # ── Filtros ──────────────────────────────────────────────────
        filters = ctk.CTkFrame(self, fg_color="#1a1a2e")
        filters.pack(fill="x", padx=12, pady=(12, 4))

        ctk.CTkLabel(filters, text="Sacramento:").pack(side="left", padx=(8, 4))
        self._tabla_var = ctk.StringVar(value="Bautismos")
        ctk.CTkComboBox(filters, variable=self._tabla_var,
                        values=list(TABLAS.keys()), width=140,
                        command=self._on_table_change).pack(side="left", padx=(0, 12))

        ctk.CTkLabel(filters, text="Año:").pack(side="left", padx=(0, 4))
        self._year_var = ctk.StringVar(value=str(datetime.datetime.now().year))
        self._year_combo = ctk.CTkComboBox(filters, variable=self._year_var,
                                           values=["Todos"], width=90)
        self._year_combo.pack(side="left", padx=(0, 12))

        ctk.CTkLabel(filters, text="Mes:").pack(side="left", padx=(0, 4))
        self._mes_var = ctk.StringVar(value="Todos")
        ctk.CTkComboBox(filters, variable=self._mes_var,
                        values=MESES, width=130).pack(side="left", padx=(0, 12))

        ctk.CTkButton(filters, text="Buscar", width=80,
                      command=self._run_query).pack(side="left", padx=(0, 8))

        self._count_lbl = ctk.CTkLabel(filters, text="")
        self._count_lbl.pack(side="left")

        # ── Treeview ─────────────────────────────────────────────────
        tree_frame = tk.Frame(self, bg="#0f0f1a")
        tree_frame.pack(fill="both", expand=True, padx=12, pady=4)

        self._tree = ttk.Treeview(
            tree_frame,
            columns=[],
            show="headings",
            selectmode="browse",
            style="NsdpTree.Treeview",
        )

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",
                            command=self._tree.yview,
                            style="NsdpTree.Vertical.TScrollbar")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal",
                            command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self._tree.pack(side="left", fill="both", expand=True)

        # ── Exportar ─────────────────────────────────────────────────
        export = ctk.CTkFrame(self, fg_color="transparent")
        export.pack(fill="x", padx=12, pady=(4, 12))

        ctk.CTkButton(export, text="Exportar a PDF",
                      fg_color="#f97316", text_color="black",
                      command=self._export_pdf).pack(side="left", padx=(0, 8))
        ctk.CTkButton(export, text="Exportar a Excel",
                      fg_color="#4ade80", text_color="black",
                      command=self._export_excel).pack(side="left")

        self._status = ctk.CTkLabel(export, text="")
        self._status.pack(side="left", padx=12)

        # Carga inicial
        self._on_table_change("Bautismos")
        self._run_query()

    # ------------------------------------------------------------------
    def _on_table_change(self, label: str):
        self._table = TABLAS.get(label, "bautismos")
        year_col = YEAR_COL[self._table]
        try:
            with db() as conn:
                rows = conn.execute(
                    f"SELECT DISTINCT {year_col} FROM {self._table} "
                    f"WHERE {year_col} IS NOT NULL ORDER BY {year_col}"
                ).fetchall()
            years = ["Todos"] + [str(r[0]) for r in rows]
        except Exception:
            years = ["Todos"]
        self._year_combo.configure(values=years)
        current_year = str(datetime.datetime.now().year)
        self._year_var.set(current_year if current_year in years else "Todos")

    def _run_query(self):
        table = self._table
        year = self._year_var.get()
        mes = self._mes_var.get()
        year_col = YEAR_COL[table]
        mes_col = MES_COL[table]
        cols = REPORT_COLS[table]

        conditions, params = [], []
        if year and year != "Todos":
            conditions.append(f"{year_col} = ?")
            params.append(year)
        if mes and mes != "Todos":
            conditions.append(f"UPPER({mes_col}) = ?")
            params.append(mes.upper())

        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        select = ", ".join(cols)
        has_parroco_col = "parroco" in cols
        order = f"parroco, {year_col}, folio" if has_parroco_col else f"{year_col}, folio"
        with db() as conn:
            rows = conn.execute(
                f"SELECT {select} FROM {table} {where} ORDER BY {order}",
                params,
            ).fetchall()

        self._results = [dict(zip(cols, r)) for r in rows]
        self._count_lbl.configure(text=f"{len(self._results)} registros")
        self._render_table(cols)

    def _render_table(self, cols: list):
        # Reconfigurar columnas del treeview
        self._tree.configure(columns=cols)
        for col in cols:
            header = REPORT_HEADERS.get(col, col.replace("_", " ").title())
            anchor = "center" if col in _CENTER_COLS else "w"
            self._tree.heading(col, text=header, anchor=anchor)
            self._tree.column(col, width=80, minwidth=40, anchor=anchor, stretch=False)

        # Limpiar y poblar
        for item in self._tree.get_children():
            self._tree.delete(item)

        self._tree.tag_configure("odd",        background=self._odd_row,  foreground="#e2e8f0")
        self._tree.tag_configure("even",       background=self._even_row, foreground="#e2e8f0")
        self._tree.tag_configure("subtotal",   background="#1e2d5e",      foreground="#93c5fd",
                                 font=("Segoe UI", 10, "bold"))
        self._tree.tag_configure("grandtotal", background="#2d1800",      foreground="#fbbf24",
                                 font=("Segoe UI", 10, "bold"))

        for i, row in enumerate(self._results[:2000]):
            values = [str(row.get(col) or "") for col in cols]
            tag = "odd" if i % 2 else "even"
            self._tree.insert("", "end", values=values, tags=(tag,))

        self._autosize_columns(cols)

        # ── Totales al final ─────────────────────────────────────────
        if not self._results:
            return

        has_parroco = "parroco" in cols
        if has_parroco:
            from collections import Counter
            counts = Counter(r.get("parroco") or "Sin párroco" for r in self._results)
            self._tree.insert("", "end", values=[""] * len(cols))
            for nombre, n in sorted(counts.items(), key=lambda x: (x[0] or "").upper()):
                vals = [""] * len(cols)
                vals[0] = f"  {nombre}"
                vals[-1] = f"{n} registros"
                self._tree.insert("", "end", values=vals, tags=("subtotal",))

        total_vals = [""] * len(cols)
        total_vals[0] = "TOTAL GENERAL"
        total_vals[-1] = f"{len(self._results)} registros"
        self._tree.insert("", "end", values=total_vals, tags=("grandtotal",))

    def _autosize_columns(self, cols: list):
        data_font = tkfont.Font(family="Segoe UI", size=10)
        head_font = tkfont.Font(family="Segoe UI", size=10, weight="bold")
        PAD = 20
        MIN = 40
        MAX = 420

        col_widths = {}
        for col in cols:
            header = REPORT_HEADERS.get(col, col.replace("_", " ").title())
            col_widths[col] = head_font.measure(header) + PAD

        for row in self._results:
            for col in cols:
                val = str(row.get(col) or "")
                w = data_font.measure(val) + PAD
                if w > col_widths[col]:
                    col_widths[col] = w

        # Longitud máxima en todos los registros de la BD (filtro actual)
        year_col = YEAR_COL[self._table]
        mes_col = MES_COL[self._table]
        year = self._year_var.get()
        mes = self._mes_var.get()
        conditions, params = [], []
        if year and year != "Todos":
            conditions.append(f"{year_col} = ?")
            params.append(year)
        if mes and mes != "Todos":
            conditions.append(f"UPPER({mes_col}) = ?")
            params.append(mes.upper())
        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

        length_exprs = ", ".join(f"MAX(LENGTH(CAST({c} AS TEXT)))" for c in cols)
        try:
            with db() as conn:
                maxlens = conn.execute(
                    f"SELECT {length_exprs} FROM {self._table} {where}", params
                ).fetchone()
            avg_char = data_font.measure("n")
            for i, col in enumerate(cols):
                estimated = avg_char * (maxlens[i] or 0) + PAD
                if estimated > col_widths[col]:
                    col_widths[col] = estimated
        except Exception:
            pass

        for col in cols:
            w = max(MIN, min(MAX, col_widths[col]))
            self._tree.column(col, width=w, minwidth=MIN)

    # ------------------------------------------------------------------
    # Exportaciones
    # ------------------------------------------------------------------
    def _export_pdf(self):
        if not self._results:
            self._status.configure(text="Sin datos para exportar.")
            return
        threading.Thread(target=self._do_export_pdf, daemon=True).start()

    def _do_export_pdf(self):
        import io
        from itertools import groupby
        from collections import Counter as _Counter
        from reportlab.pdfgen import canvas as rl_canvas
        from reportlab.lib.pagesizes import LETTER, landscape
        from reportlab.lib import colors
        from app.core.iglesia import load as _load_iglesia
        from app.utils.config import ASSETS_DIR as _ASSETS_DIR

        tabla_label = self._tabla_var.get()
        all_cols    = REPORT_COLS[self._table]
        has_parroco = "parroco" in all_cols
        cols = [c for c in all_cols if c != "parroco"] if has_parroco else all_cols

        out = Path(tempfile.gettempdir()) / f"reporte_{self._table}.pdf"
        pw, ph   = landscape(LETTER)
        MARGIN   = 40
        USABLE   = pw - 2 * MARGIN
        ROW_H    = 14
        HDR_H    = 18
        GRP_H    = 16
        LINE_H   = 10
        LOGO_H   = 50
        TBAR_H   = 22   # título del sacramento

        # Altura fija del encabezado completo:
        # logo(50) + sep(6) + gap(4) + barra_titulo(22) + gap_col(18) + col_hdr(18) + gap(4) = 122
        DATA_START_Y = ph - MARGIN - LOGO_H - 6 - 4 - TBAR_H - 18 - HDR_H - 4

        year   = self._year_var.get()
        mes    = self._mes_var.get()
        titulo = f"Reporte de {tabla_label}"
        if year != "Todos":
            titulo += f" — {year}"
        if mes != "Todos":
            titulo += f" / {mes}"

        _cfg           = _load_iglesia()
        nombre_iglesia = _cfg.get("nombre", "")
        direccion      = _cfg.get("direccion", "")
        ciudad         = _cfg.get("ciudad", "")
        addr           = "  ·  ".join(p for p in [direccion, ciudad] if p)
        parroco_cfg    = _cfg.get("parroco_actual", "")
        logo_rep       = _cfg.get("logo_reporte_file") or _cfg.get("logo_file")
        logo_path      = _ASSETS_DIR / logo_rep if logo_rep else None

        # ── Medir anchos de columna ────────────────────────────────────
        _buf = io.BytesIO()
        _mc  = rl_canvas.Canvas(_buf, pagesize=landscape(LETTER))

        desired: dict[str, float] = {}
        for col in cols:
            header = REPORT_HEADERS.get(col, col.replace("_", " ").title()).upper()
            w = _mc.stringWidth(header, "Helvetica-Bold", 8) + 14
            for row in self._results:
                w_val = _mc.stringWidth(str(row.get(col) or ""), "Helvetica", 8) + 10
                if w_val > w:
                    w = w_val
            desired[col] = w

        total_desired = sum(desired.values()) or 1.0
        scale         = min(1.0, USABLE / total_desired)
        col_widths: dict[str, float] = {col: desired[col] * scale for col in cols}

        col_x: dict[str, float] = {}
        x = MARGIN + 2
        for col in cols:
            col_x[col] = x
            x += col_widths[col]

        # ── Funciones de texto ─────────────────────────────────────────
        def fit_text(text: str, fn: str, fs: float, max_w: float) -> str:
            if _mc.stringWidth(text, fn, fs) <= max_w:
                return text
            while text and _mc.stringWidth(text + "…", fn, fs) > max_w:
                text = text[:-1]
            return (text + "…") if text else ""

        def wrap_text(text: str, fn: str, fs: float, max_w: float) -> list:
            if not text or _mc.stringWidth(text, fn, fs) <= max_w:
                return [text or ""]
            words = text.split()
            lines, current = [], ""
            for word in words:
                test = (current + " " + word).strip()
                if _mc.stringWidth(test, fn, fs) <= max_w:
                    current = test
                else:
                    if current:
                        lines.append(current)
                    if _mc.stringWidth(word, fn, fs) > max_w:
                        w = word
                        while w and _mc.stringWidth(w + "…", fn, fs) > max_w:
                            w = w[:-1]
                        current = (w + "…") if w else ""
                    else:
                        current = word
            if current:
                lines.append(current)
            return lines or [""]

        # ── Pre-calcular filas (evita recalcular durante simulación y render) ──
        def make_row_data(row: dict) -> tuple:
            cl = {col: wrap_text(str(row.get(col) or ""), "Helvetica", 8,
                                 col_widths[col] - 4) for col in cols}
            ml = max(len(v) for v in cl.values())
            return (row, cl, ml, max(ROW_H, LINE_H * ml + 4))

        if has_parroco:
            sorted_rows = sorted(self._results,
                                 key=lambda r: (r.get("parroco") or "").upper())
            grouped = [
                (pname, [make_row_data(r) for r in grp])
                for pname, grp in groupby(
                    sorted_rows, key=lambda r: r.get("parroco") or "Sin párroco"
                )
            ]
        else:
            all_rows_data = [make_row_data(r) for r in self._results]

        # ── Canvas ────────────────────────────────────────────────────
        c = rl_canvas.Canvas(str(out), pagesize=landscape(LETTER))

        # ── Funciones de dibujo (cierran sobre c) ─────────────────────
        def draw_col_header(y: float):
            c.setFillColor(colors.HexColor("#1a1a2e"))
            c.rect(MARGIN, y - 4, USABLE, HDR_H, fill=1, stroke=0)
            c.setFillColor(colors.white)
            c.setFont("Helvetica-Bold", 8)
            for col in cols:
                label = REPORT_HEADERS.get(col, col.replace("_", " ").title()).upper()
                c.drawString(col_x[col], y + 2,
                             fit_text(label, "Helvetica-Bold", 8, col_widths[col] - 4))
            c.setFont("Helvetica", 8)

        def draw_group_header(y: float, nombre: str, n: int):
            c.setFillColor(colors.HexColor("#2d3a6e"))
            c.rect(MARGIN, y - 3, USABLE, GRP_H, fill=1, stroke=0)
            c.setFillColor(colors.HexColor("#93c5fd"))
            c.setFont("Helvetica-Bold", 9)
            c.drawString(MARGIN + 6, y + 2, f"Párroco: {nombre}  ({n} registros)")
            c.setFont("Helvetica", 8)

        def draw_full_header() -> float:
            """Dibuja encabezado completo (logo + iglesia + título + col_hdr). Devuelve DATA_START_Y."""
            LOGO_Y = ph - MARGIN - LOGO_H
            TEXT_X = MARGIN + LOGO_H + 8

            if logo_path and logo_path.exists():
                try:
                    c.drawImage(str(logo_path), MARGIN, LOGO_Y,
                                width=LOGO_H, height=LOGO_H,
                                preserveAspectRatio=True, mask="auto")
                except Exception:
                    pass

            if nombre_iglesia:
                c.setFont("Helvetica-Bold", 10)
                c.setFillColor(colors.HexColor("#2a2a6a"))
                c.drawString(TEXT_X, ph - MARGIN - 14, nombre_iglesia)
            if addr:
                c.setFont("Helvetica", 8)
                c.setFillColor(colors.HexColor("#444444"))
                c.drawString(TEXT_X, ph - MARGIN - 28, addr)
            if parroco_cfg:
                c.setFont("Helvetica-Oblique", 8)
                c.setFillColor(colors.HexColor("#444444"))
                c.drawString(TEXT_X, ph - MARGIN - 42, parroco_cfg)

            sep1_y = LOGO_Y - 6
            c.setStrokeColor(colors.HexColor("#4a4a8a"))
            c.setLineWidth(0.5)
            c.line(MARGIN, sep1_y, pw - MARGIN, sep1_y)

            tbar_y = sep1_y - 4 - TBAR_H
            c.setFillColor(colors.HexColor("#2d3a6e"))
            c.rect(MARGIN, tbar_y, USABLE, TBAR_H, fill=1, stroke=0)
            c.setFillColor(colors.white)
            c.setFont("Helvetica-Bold", 11)
            c.drawCentredString(pw / 2, tbar_y + 6, titulo.upper())
            c.setFont("Helvetica", 8)
            c.drawRightString(pw - MARGIN, tbar_y + 6,
                              f"Total: {len(self._results)} registros")

            y = tbar_y - 18
            draw_col_header(y)
            return y - HDR_H - 4   # == DATA_START_Y

        def draw_footer(page: int, total: int):
            c.setFont("Helvetica", 8)
            c.setFillColor(colors.HexColor("#888888"))
            c.drawCentredString(pw / 2, 18, f"Página {page} de {total}")

        # ── Simular páginas para saber el total antes de renderizar ───
        def simulate_pages() -> int:
            page = 1
            y = DATA_START_Y

            if has_parroco:
                for _, grupo_data in grouped:
                    if y < GRP_H + ROW_H + 44:
                        page += 1
                        y = DATA_START_Y
                    y -= GRP_H + 2
                    for _, _, _, rh in grupo_data:
                        if y < 44 + (rh - ROW_H):
                            page += 1
                            y = DATA_START_Y
                        y -= rh
                    y -= 4
            else:
                for _, _, _, rh in all_rows_data:
                    if y < 44 + (rh - ROW_H):
                        page += 1
                        y = DATA_START_Y
                    y -= rh

            # sección totales
            y -= 8
            if y < 50:
                page += 1
                y = DATA_START_Y
            y -= 14 + HDR_H + 4
            if has_parroco:
                for _ in sorted(
                    _Counter(r.get("parroco") or "Sin párroco"
                             for r in self._results).items()
                ):
                    if y < 30:
                        page += 1
                        y = DATA_START_Y
                    y -= ROW_H
                if y < 30:
                    page += 1

            return page

        total_pages = simulate_pages()

        # ── Renderizar ─────────────────────────────────────────────────
        current_page = 1
        y = draw_full_header()

        if has_parroco:
            for parroco_nombre, grupo_data in grouped:
                if y < GRP_H + ROW_H + 44:
                    draw_footer(current_page, total_pages)
                    c.showPage()
                    current_page += 1
                    y = draw_full_header()

                draw_group_header(y, parroco_nombre, len(grupo_data))
                y -= GRP_H + 2

                for i, (row, cell_lines, max_ln, row_h) in enumerate(grupo_data):
                    if y < 44 + (row_h - ROW_H):
                        draw_footer(current_page, total_pages)
                        c.showPage()
                        current_page += 1
                        y = draw_full_header()

                    bg = colors.HexColor("#eef2ff") if i % 2 == 0 else colors.white
                    c.setFillColor(bg)
                    c.rect(MARGIN, y - 3 - LINE_H * (max_ln - 1), USABLE, row_h,
                           fill=1, stroke=0)
                    c.setFillColor(colors.black)
                    for col in cols:
                        for li, line in enumerate(cell_lines[col]):
                            c.drawString(col_x[col], y + 1 - li * LINE_H, line)
                    y -= row_h

                y -= 4
        else:
            for i, (row, cell_lines, max_ln, row_h) in enumerate(all_rows_data):
                if y < 44 + (row_h - ROW_H):
                    draw_footer(current_page, total_pages)
                    c.showPage()
                    current_page += 1
                    y = draw_full_header()

                bg = colors.HexColor("#eef2ff") if i % 2 == 0 else colors.white
                c.setFillColor(bg)
                c.rect(MARGIN, y - 3 - LINE_H * (max_ln - 1), USABLE, row_h,
                       fill=1, stroke=0)
                c.setFillColor(colors.black)
                for col in cols:
                    for li, line in enumerate(cell_lines[col]):
                        c.drawString(col_x[col], y + 1 - li * LINE_H, line)
                y -= row_h

        # ── Sección de totales finales ─────────────────────────────────
        y -= 8
        if y < 50:
            draw_footer(current_page, total_pages)
            c.showPage()
            current_page += 1
            y = draw_full_header()

        c.setStrokeColor(colors.HexColor("#4a5568"))
        c.line(MARGIN, y, MARGIN + USABLE, y)
        y -= 14

        c.setFillColor(colors.HexColor("#1a1a2e"))
        c.rect(MARGIN, y - 4, USABLE, HDR_H, fill=1, stroke=0)
        c.setFillColor(colors.HexColor("#fbbf24"))
        c.setFont("Helvetica-Bold", 9)
        c.drawString(MARGIN + 6, y + 2, "RESUMEN DE TOTALES")
        y -= HDR_H + 4

        if has_parroco:
            parroco_counts = _Counter(
                r.get("parroco") or "Sin párroco" for r in self._results
            )
            c.setFont("Helvetica", 8)
            for i, (nombre, n) in enumerate(
                sorted(parroco_counts.items(), key=lambda x: (x[0] or "").upper())
            ):
                if y < 30:
                    draw_footer(current_page, total_pages)
                    c.showPage()
                    current_page += 1
                    y = draw_full_header()
                    c.setFont("Helvetica", 8)
                bg = colors.HexColor("#eef2ff") if i % 2 == 0 else colors.white
                c.setFillColor(bg)
                c.rect(MARGIN, y - 3, USABLE, ROW_H, fill=1, stroke=0)
                c.setFillColor(colors.black)
                c.drawString(MARGIN + 6, y + 1, f"  {nombre}")
                c.drawRightString(MARGIN + USABLE - 6, y + 1, f"{n} registros")
                y -= ROW_H
            if y < 30:
                draw_footer(current_page, total_pages)
                c.showPage()
                current_page += 1
                y = draw_full_header()

        c.setFillColor(colors.HexColor("#2d1800"))
        c.rect(MARGIN, y - 3, USABLE, ROW_H + 2, fill=1, stroke=0)
        c.setFillColor(colors.HexColor("#fbbf24"))
        c.setFont("Helvetica-Bold", 9)
        c.drawString(MARGIN + 6, y + 1, "TOTAL GENERAL")
        c.drawRightString(MARGIN + USABLE - 6, y + 1, f"{len(self._results)} registros")

        draw_footer(current_page, total_pages)
        c.save()
        self.after(0, self._open_and_notify, out, "PDF")

    def _export_excel(self):
        if not self._results:
            self._status.configure(text="Sin datos para exportar.")
            return
        threading.Thread(target=self._do_export_excel, daemon=True).start()

    def _do_export_excel(self):
        import openpyxl
        from itertools import groupby
        from openpyxl.styles import Font, PatternFill, Alignment

        all_cols    = REPORT_COLS[self._table]
        has_parroco = "parroco" in all_cols
        # Igual que el PDF: si hay párroco, se omite de las columnas de fila
        cols = [c for c in all_cols if c != "parroco"] if has_parroco else all_cols

        out = Path(tempfile.gettempdir()) / f"reporte_{self._table}.xlsx"
        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = self._tabla_var.get()[:31]

        # ── Estilos ───────────────────────────────────────────────────
        hdr_fill  = PatternFill("solid", fgColor="1A1A2E")
        hdr_font  = Font(bold=True, color="FFFFFF", size=10)
        grp_fill  = PatternFill("solid", fgColor="2D3A6E")
        grp_font  = Font(bold=True, color="93C5FD", size=10)
        alt_fill  = PatternFill("solid", fgColor="EEF2FF")
        sub_fill  = PatternFill("solid", fgColor="1E2D5E")
        sub_font  = Font(bold=True, color="93C5FD", size=10)
        tot_fill  = PatternFill("solid", fgColor="2D1800")
        tot_font  = Font(bold=True, color="FBBF24", size=11)

        def style_row(row_idx, fill, font=None):
            for j in range(1, len(cols) + 1):
                cell = ws.cell(row=row_idx, column=j)
                cell.fill = fill
                if font:
                    cell.font = font

        # ── Encabezado de columnas ─────────────────────────────────────
        for j, col in enumerate(cols, 1):
            header = REPORT_HEADERS.get(col, col.replace("_", " ").title())
            cell = ws.cell(row=1, column=j, value=header.upper())
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = 18

        # ── Datos (con corte de grupo por párroco) ─────────────────────
        if has_parroco:
            sorted_rows = sorted(
                self._results,
                key=lambda r: (r.get("parroco") or "").upper(),
            )
            groups = groupby(sorted_rows, key=lambda r: r.get("parroco") or "Sin párroco")

            for parroco_nombre, grupo_iter in groups:
                grupo = list(grupo_iter)

                # Fila de encabezado de grupo
                grp_vals = [""] * len(cols)
                grp_vals[0] = f"Párroco: {parroco_nombre}  ({len(grupo)} registros)"
                ws.append(grp_vals)
                style_row(ws.max_row, grp_fill, grp_font)

                for i, row in enumerate(grupo):
                    data_vals = [row.get(col) for col in cols]
                    ws.append(data_vals)
                    fill = alt_fill if i % 2 == 0 else PatternFill()
                    style_row(ws.max_row, fill)
        else:
            for i, row in enumerate(self._results):
                data_vals = [row.get(col) for col in cols]
                ws.append(data_vals)
                fill = alt_fill if i % 2 == 0 else PatternFill()
                style_row(ws.max_row, fill)

        # ── Totales finales ───────────────────────────────────────────
        ws.append([""] * len(cols))   # separador

        if has_parroco:
            from collections import Counter
            counts = Counter(r.get("parroco") or "Sin párroco" for r in self._results)
            for nombre, n in sorted(counts.items(), key=lambda x: (x[0] or "").upper()):
                sub_vals = [""] * len(cols)
                sub_vals[0] = f"Subtotal — {nombre}"
                sub_vals[-1] = f"{n} registros"
                ws.append(sub_vals)
                style_row(ws.max_row, sub_fill, sub_font)

        tot_vals = [""] * len(cols)
        tot_vals[0] = "TOTAL GENERAL"
        tot_vals[-1] = f"{len(self._results)} registros"
        ws.append(tot_vals)
        style_row(ws.max_row, tot_fill, tot_font)

        wb.save(out)
        self.after(0, self._open_and_notify, out, "Excel")

    def _open_and_notify(self, path: Path, fmt: str):
        import os
        try:
            import win32api
            win32api.ShellExecute(0, "open", str(path), None, ".", 1)
        except Exception:
            os.startfile(str(path))
        self._status.configure(text=f"{fmt} generado: {path.name}")
